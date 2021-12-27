import datetime
from openpyxl import load_workbook


def get_absences(props):
    try:
        assert props['absenceFile'] != None, 'Hour tracking system Absence Excel file not set'
    except Exception as ex:
        print("Properties file not complete:", repr(ex))
        exit()

    wb = load_workbook(props['absenceFile'])

    absences = []
    for row in wb.active.iter_rows(min_row=2):
        external_id, start_date, end_date, approval_type = row
        absence = {
            'externalId': external_id.value,
            'startDate': start_date.value.strftime('%Y-%m-%d')
        }
        if end_date.value:
            absence['endDate'] = end_date.value.strftime('%Y-%m-%d')
        if approval_type.value:
            absence['approvalType'] = approval_type.value
        absences.append(absence)

    return absences
